import os
import re
import csv
import openpyxl
import argparse

def main():
    parser = argparse.ArgumentParser(description="Cập nhật thông tin hàng loạt (học phí, tài khoản ngân hàng, liên hệ, hoặc thay thế văn bản bất kỳ) cho chatbot/website của các khóa học.")
    parser.add_argument("--domain", type=str, help="Mã domain khóa học cụ thể cần cập nhật (ví dụ: 03_domain_oplw). Nếu bỏ qua sẽ áp dụng cho tất cả.")
    parser.add_argument("--summary", type=str, default="/root/Second_Brain/01_chuong_trinh_dao_tao/00_bang_tong_hop_thong_tin_cac_chuong_trinh.md", help="Đường dẫn đến file summary để parse học phí.")
    
    # 1. Chức năng thay thế văn bản bất kỳ
    parser.add_argument("--search", type=str, help="Văn bản cũ cần tìm để thay thế (dùng cho thay thế chung).")
    parser.add_argument("--replace", type=str, help="Văn bản mới thay thế vào (dùng kèm với --search).")
    
    # 2. Chức năng cập nhật học phí theo mẫu tư vấn
    parser.add_argument("--tuition", action="store_true", help="Kích hoạt cập nhật học phí theo mẫu tư vấn.")
    parser.add_argument("--price-override", type=str, help="Ghi đè giá học phí dạng K (ví dụ: 1.480K). Nếu không truyền sẽ lấy từ summary.")
    parser.add_argument("--template-override", type=str, help="Ghi đè kịch bản trả lời học phí của AI. Sử dụng placeholder {price_K}.")
    
    # 3. Phím tắt cập nhật thông tin phổ biến (Số TK, Ngân hàng, Hotline, Zalo,...)
    parser.add_argument("--bank-acc", type=str, help="Cập nhật số tài khoản ngân hàng mới (thay thế cho TK cũ '19034464432011').")
    parser.add_argument("--bank-name", type=str, help="Cập nhật tên ngân hàng mới (thay thế cho 'Techcombank').")
    parser.add_argument("--bank-owner", type=str, help="Cập nhật chủ tài khoản mới (thay thế cho 'Lê Thị Thu Nhi').")
    parser.add_argument("--zalo", type=str, help="Cập nhật số Zalo kích hoạt khóa học mới (thay thế cho '0854838394').")
    parser.add_argument("--hotline", type=str, help="Cập nhật hotline công ty mới (thay thế cho '0914972102').")
    parser.add_argument("--email", type=str, help="Cập nhật email hỗ trợ mới (thay thế cho 'contact@anhlaptrinh.com').")

    args = parser.parse_args()
    base_dir = "/root/Second_Brain/01_chuong_trinh_dao_tao"

    # Mặc định của các trường thông tin cũ để tìm kiếm thay thế nhanh
    OLD_BANK_ACC = "19034464432011"
    OLD_BANK_NAME = "Techcombank"
    OLD_BANK_OWNER = "Lê Thị Thu Nhi"
    OLD_ZALO = "0854838394"
    OLD_HOTLINE = "0914972102"
    OLD_EMAIL = "contact@anhlaptrinh.com"

    # Tạo danh sách các cặp tìm kiếm - thay thế
    replacements = []

    # Thêm các phím tắt cập nhật thông tin nếu được truyền vào
    if args.bank_acc:
        replacements.append((OLD_BANK_ACC, args.bank_acc))
    if args.bank_name:
        replacements.append((OLD_BANK_NAME, args.bank_name))
    if args.bank_owner:
        replacements.append((OLD_BANK_OWNER, args.bank_owner))
    if args.zalo:
        replacements.append((OLD_ZALO, args.zalo))
    if args.hotline:
        replacements.append((OLD_HOTLINE, args.hotline))
    if args.email:
        replacements.append((OLD_EMAIL, args.email))
        # Dự phòng trường hợp có đuôi .vn thay vì .com
        replacements.append(("contact@anhlaptrinh.vn", args.email))

    # Thêm cặp tìm kiếm - thay thế tự do nếu có
    if args.search is not None and args.replace is not None:
        replacements.append((args.search, args.replace))

    # Nếu không có hành động nào được chọn
    if not replacements and not args.tuition:
        print("Không có tham số cập nhật nào được truyền. Vui lòng truyền các tham số như --search/--replace, --tuition, hoặc các phím tắt như --bank-acc, --zalo,...")
        return

    # Parse giá từ summary nếu cần cập nhật học phí
    price_map = {}
    if args.tuition and os.path.exists(args.summary):
        with open(args.summary, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines:
            if line.strip().startswith("|") and "`" in line:
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 5:
                    domain_match = re.search(r'`([^`]+)`', parts[1])
                    if domain_match:
                        domain_code = domain_match.group(1)
                        tuition_raw = parts[4]
                        match = re.search(r'(\d+)\.(\d{3})\.(\d{3})', tuition_raw)
                        if match:
                            price_K = f"{match.group(1)}.{match.group(2)}K"
                            price_long = f"{match.group(1)}.{match.group(2)}.{match.group(3)}đ"
                        else:
                            price_K = "1.460K"
                            price_long = "1.460.000đ"
                        price_map[domain_code] = {"K": price_K, "long": price_long}

    # Ghi đè giá thủ công cho học phí nếu được truyền
    if args.price_override:
        for k in price_map:
            price_map[k]["K"] = args.price_override

    # Template học phí mới
    default_template = (
        "Phí trọn gói là {price_K} anh nhé, giờ anh đăng ký là vào học và tạo trợ lý ngay và luôn ạ.  \n\n"
        "Anh sẽ sở hữu cho bản thân những trợ lý phục vụ công việc đặc thù của chính anh,  giúp anh x3 hiệu quả công việc."
    )
    ans_template = args.template_override if args.template_override else default_template

    def get_new_answer(price_K):
        return ans_template.replace("{price_K}", price_K)

    def get_new_blockquote(price_K):
        raw_ans = get_new_answer(price_K)
        lines = raw_ans.splitlines()
        bq_lines = []
        for line in lines:
            if line.strip() == "":
                bq_lines.append(">")
            else:
                bq_lines.append(f"> {line}")
        return "\n".join(bq_lines)

    # Hàm thực hiện thay thế văn bản bất kỳ trong file CSV
    def apply_text_replacements_csv(csv_path, pairs):
        updated_rows = []
        changed = 0
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if not row:
                    continue
                new_row = []
                for cell in row:
                    new_cell = cell
                    for old_str, new_str in pairs:
                        if old_str in new_cell:
                            new_cell = new_cell.replace(old_str, new_str)
                            changed += 1
                    new_row.append(new_cell)
                updated_rows.append(new_row)
        
        if changed > 0:
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(updated_rows)
            print(f"  CSV: Đã thay thế {changed} cụm từ.")
        return changed

    # Hàm thực hiện thay thế văn bản bất kỳ trong file XLSX
    def apply_text_replacements_xlsx(xlsx_path, pairs):
        wb = openpyxl.load_workbook(xlsx_path)
        sheet_name = "du_lieu_nap_ai"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        changed = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_val = cell.value
                    for old_str, new_str in pairs:
                        if old_str in new_val:
                            new_val = new_val.replace(old_str, new_str)
                            changed += 1
                    cell.value = new_val
        
        if changed > 0:
            wb.save(xlsx_path)
            print(f"  XLSX: Đã thay thế {changed} cụm từ trong sheet '{sheet_name}'.")
        return changed

    # Hàm thực hiện thay thế văn bản bất kỳ trong file Markdown
    def apply_text_replacements_md(md_path, pairs):
        with open(md_path, "r", encoding="utf-8") as f:
            content = f.read()
        
        new_content = content
        changed = 0
        for old_str, new_str in pairs:
            if old_str in new_content:
                occurrences = new_content.count(old_str)
                new_content = new_content.replace(old_str, new_str)
                changed += occurrences
                print(f"  MD: Đã thay thế {occurrences} lần '{old_str[:30]}...' -> '{new_str[:30]}...'.")
                
        if changed > 0:
            with open(md_path, "w", encoding="utf-8") as f:
                f.write(new_content)
        return changed

    # Hàm xử lý cập nhật học phí trong CSV
    def update_tuition_csv(csv_path, price_K):
        updated_rows = []
        changed = 0
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
            updated_rows.append(header)
            for row in reader:
                if not row:
                    continue
                question, answer = row[0], row[1]
                clean_q = question.strip().lower().rstrip("?").strip()
                if clean_q in ["học phí bao nhiêu", "học phí khóa này bao nhiêu vậy em"]:
                    new_ans = get_new_answer(price_K)
                    if answer != new_ans:
                        answer = new_ans
                        changed += 1
                updated_rows.append([question, answer])
        
        found = any(row[0].strip().lower().rstrip("?").strip() == "học phí bao nhiêu" for row in updated_rows[1:])
        if not found:
            updated_rows.append(["Học phí bao nhiêu?", get_new_answer(price_K)])
            changed += 1
            print("  CSV: Đã thêm dòng câu hỏi học phí mới.")
        
        if changed > 0:
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(updated_rows)
            print(f"  CSV: Đã cập nhật học phí cho {changed} dòng.")
        return changed

    # Hàm xử lý cập nhật học phí trong XLSX
    def update_tuition_xlsx(xlsx_path, price_K):
        wb = openpyxl.load_workbook(xlsx_path)
        sheet_name = "du_lieu_nap_ai"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        changed = 0
        found = False
        new_ans = get_new_answer(price_K)
        for row_idx in range(2, sheet.max_row + 1):
            q_cell = sheet.cell(row=row_idx, column=1)
            a_cell = sheet.cell(row=row_idx, column=2)
            if q_cell.value:
                clean_q = q_cell.value.strip().lower().rstrip("?").strip()
                if clean_q in ["học phí bao nhiêu", "học phí khóa này bao nhiêu vậy em"]:
                    found = True
                    if a_cell.value != new_ans:
                        a_cell.value = new_ans
                        changed += 1
        
        if not found:
            sheet.append(["Học phí bao nhiêu?", new_ans])
            changed += 1
            print("  XLSX: Đã thêm dòng câu hỏi học phí mới.")
        
        if changed > 0:
            wb.save(xlsx_path)
            print(f"  XLSX: Đã cập nhật học phí cho {changed} dòng trong sheet '{sheet_name}'.")
        return changed

    # Hàm xử lý cập nhật học phí trong MD
    def update_tuition_markdown(md_path, price_K):
        with open(md_path, "r", encoding="utf-8") as f:
            content = f.read()
        
        sections = re.split(r'^(#+\s+.*)$', content, flags=re.MULTILINE)
        changed = 0
        header_indices = [i for i, sec in enumerate(sections) if sec.strip().startswith("#")]
        
        tuition_headers = [
            "học phí bao nhiêu",
            "khách hỏi học phí",
            "khách hỏi giá",
            "kịch bản phản hồi chi tiết",
            "khách hỏi học phí/lịch học",
            "khi khách hỏi học phí"
        ]
        
        found_header = False
        for idx in header_indices:
            header = sections[idx].strip().lower()
            if any(h in header for h in tuition_headers):
                found_header = True
                if idx + 1 < len(sections):
                    sec_content = sections[idx + 1]
                    lines = sec_content.splitlines(keepends=True)
                    new_lines = []
                    replaced_blockquote = False
                    replaced_ai = False
                    replaced_tra_loi = False
                    
                    for line in lines:
                        stripped = line.strip()
                        if stripped.startswith(">"):
                            if not replaced_blockquote:
                                new_lines.append(get_new_blockquote(price_K) + "\n")
                                replaced_blockquote = True
                            continue
                        
                        match_ai = re.match(r'^(\*?\*?AI\*?\*?:)\s*(.*)$', stripped, re.IGNORECASE)
                        if match_ai:
                            prefix = match_ai.group(1)
                            new_lines.append(f"{prefix} {get_new_answer(price_K)}\n")
                            replaced_ai = True
                            continue
                        
                        match_tl = re.match(r'^(\*?\*?Trả lời\*?\*?:)\s*(.*)$', stripped, re.IGNORECASE)
                        if match_tl:
                            prefix = match_tl.group(1)
                            new_lines.append(f"{prefix} {get_new_answer(price_K)}\n")
                            replaced_tra_loi = True
                            continue
                        
                        new_lines.append(line)
                    
                    if not (replaced_blockquote or replaced_ai or replaced_tra_loi):
                        cleaned_lines = [l for l in new_lines if l.strip()]
                        new_sec_content = "\n" + get_new_answer(price_K) + "\n\n"
                    else:
                        new_sec_content = "".join(new_lines)
                    
                    if sections[idx + 1] != new_sec_content:
                        sections[idx + 1] = new_sec_content
                        changed += 1
                        print(f"  MD: Đã cập nhật câu trả lời học phí dưới mục '{sections[idx].strip()}'.")
        
        new_content = "".join(sections)
        
        if not found_header:
            append_str = (
                f"\n\n## Học phí bao nhiêu?\n\n"
                f"{get_new_answer(price_K)}\n"
            )
            new_content = new_content.rstrip() + append_str
            changed += 1
            print("  MD: Đã thêm mục 'Học phí bao nhiêu?' ở cuối file.")
            
        if changed > 0:
            with open(md_path, "w", encoding="utf-8") as f:
                f.write(new_content)
        return changed

    # 4. Xác định các domain đích
    all_domains = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d)) and d.split("_")[0].isdigit()]
    all_domains.sort(key=lambda x: int(x.split("_")[0]))
    
    target_domains = []
    if args.domain:
        if args.domain in all_domains:
            target_domains = [args.domain]
        else:
            print(f"Lỗi: Không tìm thấy domain '{args.domain}' trong hệ thống.")
            return
    else:
        target_domains = all_domains

    # 5. Chạy cập nhật
    report = []
    for d in target_domains:
        d_path = os.path.join(base_dir, d)
        subdirs = [sd for sd in os.listdir(d_path) if os.path.isdir(os.path.join(d_path, sd)) and sd.startswith("01_du_lieu_website_chatbot_")]
        if not subdirs:
            continue
        
        active_subdir = subdirs[0]
        sub_path = os.path.join(d_path, active_subdir)
        
        print(f"\n======================================")
        print(f"Xử lý chương trình: {d}")
        print(f"======================================")
        
        csv_path = os.path.join(sub_path, "01_du_lieu_nap_ai_fanpage.csv")
        xlsx_path = os.path.join(sub_path, "02_du_lieu_nap_ai_fanpage.xlsx")
        md_tu_van_path = os.path.join(sub_path, "03_du_lieu_chatbot_tu_van.md")
        md_website_path = os.path.join(sub_path, "04_du_lieu_lam_website.md")
        
        files_updated = []

        # A. Cập nhật học phí theo template
        if args.tuition:
            price_K = "1.460K"
            if d in price_map:
                price_K = price_map[d]["K"]
            
            if os.path.exists(csv_path):
                c = update_tuition_csv(csv_path, price_K)
                if c > 0: files_updated.append(os.path.basename(csv_path))
            if os.path.exists(xlsx_path):
                c = update_tuition_xlsx(xlsx_path, price_K)
                if c > 0: files_updated.append(os.path.basename(xlsx_path))
            if os.path.exists(md_tu_van_path):
                c = update_tuition_markdown(md_tu_van_path, price_K)
                if c > 0: files_updated.append(os.path.basename(md_tu_van_path))
            if os.path.exists(md_website_path):
                c = update_tuition_markdown(md_website_path, price_K)
                if c > 0: files_updated.append(os.path.basename(md_website_path))

        # B. Cập nhật thay thế văn bản hàng loạt
        if replacements:
            if os.path.exists(csv_path):
                c = apply_text_replacements_csv(csv_path, replacements)
                if c > 0 and os.path.basename(csv_path) not in files_updated:
                    files_updated.append(os.path.basename(csv_path))
            if os.path.exists(xlsx_path):
                c = apply_text_replacements_xlsx(xlsx_path, replacements)
                if c > 0 and os.path.basename(xlsx_path) not in files_updated:
                    files_updated.append(os.path.basename(xlsx_path))
            if os.path.exists(md_tu_van_path):
                c = apply_text_replacements_md(md_tu_van_path, replacements)
                if c > 0 and os.path.basename(md_tu_van_path) not in files_updated:
                    files_updated.append(os.path.basename(md_tu_van_path))
            if os.path.exists(md_website_path):
                c = apply_text_replacements_md(md_website_path, replacements)
                if c > 0 and os.path.basename(md_website_path) not in files_updated:
                    files_updated.append(os.path.basename(md_website_path))

        if files_updated:
            report.append({
                "domain": d,
                "files": files_updated
            })

    if report:
        print("\n\n======================================")
        print("BÁO CÁO HOÀN THÀNH CẬP NHẬT")
        print("======================================")
        for item in report:
            print(f"Domain {item['domain']}:")
            print(f"  Các file đã cập nhật: {', '.join(item['files'])}")
    else:
        print("\nKhông có tệp nào thay đổi.")

if __name__ == "__main__":
    main()
