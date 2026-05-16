#!/usr/bin/env python3
"""Generate clean demo datasets for Anh Lap Trinh lessons."""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date, timedelta
from pathlib import Path
from random import Random


SKILL_DIR = Path(__file__).resolve().parents[1]
DEFAULT_LOGO_PATH = SKILL_DIR / "assets" / "logo.png"
DEFAULT_COMPANY_INFO = {
    "company_name": "Cty TNHH Anh Lap Trinh",
    "director": "Nguyen Van Nhuan",
    "note": "Demo data for Anh Lap Trinh training. All business records are fictional.",
}

FIRST_NAMES = ["An", "Binh", "Chi", "Dung", "Hanh", "Khoa", "Linh", "Minh", "Nam", "Phuong", "Quang", "Trang"]
LAST_NAMES = ["Nguyen", "Tran", "Le", "Pham", "Hoang", "Vo", "Dang", "Bui", "Do", "Ngo"]
CITIES = ["Ha Noi", "TP HCM", "Da Nang", "Can Tho", "Hai Phong", "Nha Trang", "Hue", "Binh Duong"]
OWNERS = ["Nhuan", "Hoa", "Tuan", "Mai", "Long", "Thao"]


def person_name(rng: Random) -> str:
    return f"{rng.choice(LAST_NAMES)} {rng.choice(FIRST_NAMES)}"


def money(value: float) -> int:
    return int(round(value / 1000) * 1000)


def pick_date(rng: Random, start: date, span_days: int) -> str:
    return (start + timedelta(days=rng.randint(0, span_days))).isoformat()


def sales_orders(rows: int, rng: Random) -> list[dict[str, object]]:
    products = [
        ("Python Automation Starter", "Course", 2500000),
        ("Excel Report Template", "Template", 350000),
        ("CRM Mini App", "Software", 1200000),
        ("AI Assistant Setup", "Service", 3000000),
        ("Google Sheets Toolkit", "Template", 450000),
    ]
    result = []
    for i in range(1, rows + 1):
        product, category, price = rng.choice(products)
        quantity = rng.randint(1, 5)
        discount = rng.choice([0, 0, 5, 10, 15])
        result.append(
            {
                "order_id": f"ORD{i:03d}",
                "order_date": pick_date(rng, date(2026, 1, 1), 120),
                "customer_id": f"CUS{rng.randint(1, max(5, rows // 3)):03d}",
                "customer_name": person_name(rng),
                "city": rng.choice(CITIES),
                "product_name": product,
                "category": category,
                "quantity": quantity,
                "unit_price": price,
                "discount_percent": discount,
                "payment_status": rng.choice(["paid", "unpaid", "partial"]),
                "delivery_status": rng.choice(["new", "processing", "delivered", "cancelled"]),
                "sales_owner": rng.choice(OWNERS),
                "total_amount": money(quantity * price * (1 - discount / 100)),
            }
        )
    return result


def office_tasks(rows: int, rng: Random) -> list[dict[str, object]]:
    departments = ["Sales", "Marketing", "Finance", "HR", "Operations", "Training"]
    tasks = ["Prepare report", "Clean spreadsheet", "Call customer", "Update CRM", "Check payment", "Send class reminder"]
    result = []
    for i in range(1, rows + 1):
        created = date(2026, 2, 1) + timedelta(days=rng.randint(0, 45))
        due = created + timedelta(days=rng.randint(1, 14))
        result.append(
            {
                "task_id": f"TSK{i:03d}",
                "created_date": created.isoformat(),
                "department": rng.choice(departments),
                "owner_name": rng.choice(OWNERS),
                "task_title": rng.choice(tasks),
                "priority": rng.choice(["low", "medium", "high"]),
                "due_date": due.isoformat(),
                "status": rng.choice(["todo", "doing", "done", "blocked"]),
                "estimated_hours": rng.choice([1, 2, 3, 4, 6, 8]),
                "actual_hours": rng.choice([0, 1, 2, 3, 4, 5, 8, 10]),
            }
        )
    return result


def leads_crm(rows: int, rng: Random) -> list[dict[str, object]]:
    sources = ["Facebook", "Zalo", "Website", "Referral", "Workshop", "YouTube"]
    interests = ["Python", "Automation", "AI Assistant", "Excel", "OpenClaw", "Vibe Coding"]
    stages = ["new", "contacted", "demo_scheduled", "won", "lost"]
    result = []
    for i in range(1, rows + 1):
        result.append(
            {
                "lead_id": f"LEAD{i:03d}",
                "created_date": pick_date(rng, date(2026, 1, 15), 90),
                "lead_name": person_name(rng),
                "company_name": f"Cong Ty {rng.choice(['An Phat', 'Minh Tam', 'Sao Viet', 'Hoa Binh', 'Tan Thanh'])}",
                "city": rng.choice(CITIES),
                "source": rng.choice(sources),
                "interest": rng.choice(interests),
                "budget_vnd": money(rng.randint(2, 30) * 1000000),
                "owner_name": rng.choice(OWNERS),
                "next_follow_up": pick_date(rng, date(2026, 5, 1), 21),
                "stage": rng.choice(stages),
                "note": rng.choice(["needs demo", "asked for price", "busy this week", "ready to join", "needs manager approval"]),
            }
        )
    return result


def course_students(rows: int, rng: Random) -> list[dict[str, object]]:
    courses = ["Python Automation", "AI Assistant", "Excel Automation", "OpenClaw", "Vibe Coding"]
    result = []
    for i in range(1, rows + 1):
        total = rng.choice([6, 8, 10, 12])
        attended = rng.randint(0, total)
        result.append(
            {
                "student_id": f"STU{i:03d}",
                "registered_date": pick_date(rng, date(2026, 1, 1), 120),
                "student_name": person_name(rng),
                "course_name": rng.choice(courses),
                "class_code": f"ALT{rng.randint(1, 9):02d}",
                "payment_status": rng.choice(["paid", "unpaid", "partial"]),
                "attendance_sessions": attended,
                "total_sessions": total,
                "homework_status": rng.choice(["not_started", "in_progress", "submitted", "reviewed"]),
                "certificate_status": "ready" if attended / total >= 0.8 else "not_ready",
                "support_owner": rng.choice(OWNERS),
            }
        )
    return result


def inventory(rows: int, rng: Random) -> list[dict[str, object]]:
    categories = ["Device", "Book", "Template", "Accessory", "Software"]
    warehouses = ["Ha Noi", "TP HCM", "Da Nang"]
    result = []
    for i in range(1, rows + 1):
        stock = rng.randint(0, 200)
        reorder = rng.choice([10, 20, 30, 50])
        result.append(
            {
                "sku": f"SKU{i:04d}",
                "product_name": f"Demo Product {i:03d}",
                "category": rng.choice(categories),
                "supplier_name": f"Supplier {rng.choice(['A', 'B', 'C', 'D'])}",
                "warehouse": rng.choice(warehouses),
                "stock_quantity": stock,
                "reorder_level": reorder,
                "unit_cost": money(rng.randint(50, 1500) * 1000),
                "last_updated": pick_date(rng, date(2026, 4, 1), 40),
                "status": "low_stock" if stock <= reorder else "ok",
            }
        )
    return result


def hr_attendance(rows: int, rng: Random) -> list[dict[str, object]]:
    departments = ["Sales", "Marketing", "Finance", "HR", "Operations", "Training"]
    titles = ["Staff", "Specialist", "Leader", "Manager"]
    result = []
    for i in range(1, rows + 1):
        late = rng.choice([False, False, False, True])
        check_in = "09:15" if late else rng.choice(["08:00", "08:15", "08:30", "08:45"])
        leave = rng.choice(["none", "none", "none", "annual_leave", "sick_leave"])
        result.append(
            {
                "employee_id": f"EMP{rng.randint(1, max(5, rows // 2)):03d}",
                "work_date": pick_date(rng, date(2026, 4, 1), 30),
                "employee_name": person_name(rng),
                "department": rng.choice(departments),
                "job_title": rng.choice(titles),
                "check_in": "" if leave != "none" else check_in,
                "check_out": "" if leave != "none" else rng.choice(["17:00", "17:30", "18:00"]),
                "work_hours": 0 if leave != "none" else rng.choice([7.5, 8, 8.5, 9]),
                "leave_status": leave,
                "approval_status": rng.choice(["approved", "pending", "rejected"]),
            }
        )
    return result


SCENARIOS = {
    "sales-orders": sales_orders,
    "office-tasks": office_tasks,
    "leads-crm": leads_crm,
    "course-students": course_students,
    "inventory": inventory,
    "hr-attendance": hr_attendance,
}


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, rows: list[dict[str, object]]) -> None:
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def write_company_csv(path: Path, company_info: dict[str, str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["field", "value"])
        for key, value in company_info.items():
            writer.writerow([key, value])


def write_company_json(path: Path, company_info: dict[str, str]) -> None:
    path.write_text(json.dumps(company_info, ensure_ascii=False, indent=2), encoding="utf-8")


def write_xlsx(path: Path, rows: list[dict[str, object]], company_info: dict[str, str] | None = None, logo_path: Path | None = None) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "demo_data"
    if company_info:
        info_ws = wb.create_sheet("company_info", 0)
        info_ws["A1"] = company_info["company_name"]
        info_ws["A1"].font = Font(bold=True, size=16)
        info_ws["A2"] = f"Director: {company_info['director']}"
        info_ws["A3"] = company_info["note"]
        info_ws.column_dimensions["A"].width = 55
        if logo_path and logo_path.exists():
            try:
                from openpyxl.drawing.image import Image

                logo = Image(str(logo_path))
                logo.width = 140
                logo.height = 70
                info_ws.add_image(logo, "C1")
            except Exception as exc:
                info_ws["A5"] = f"Logo skipped: {exc}"

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, header in enumerate(headers, start=1):
        max_len = max(len(str(header)), *(len(str(row[header])) for row in rows[:100]))
        ws.column_dimensions[get_column_letter(index)].width = min(max_len + 2, 32)
    wb.save(path)
    return True


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Anh Lap Trinh demo data.")
    parser.add_argument("--scenario", choices=sorted(SCENARIOS), required=True)
    parser.add_argument("--rows", type=int, default=50)
    parser.add_argument("--out", type=Path, default=Path("demo-data"))
    parser.add_argument("--formats", default="csv,json", help="Comma-separated: csv,json,xlsx")
    parser.add_argument("--seed", type=int, default=20260513)
    parser.add_argument("--include-company-info", action="store_true", help="Write Anh Lap Trinh company metadata.")
    parser.add_argument("--company-name", default=DEFAULT_COMPANY_INFO["company_name"])
    parser.add_argument("--director-name", default=DEFAULT_COMPANY_INFO["director"])
    parser.add_argument("--logo", type=Path, default=DEFAULT_LOGO_PATH)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.rows < 1:
        raise SystemExit("--rows must be at least 1")

    formats = {item.strip().lower() for item in args.formats.split(",") if item.strip()}
    unknown = formats - {"csv", "json", "xlsx"}
    if unknown:
        raise SystemExit(f"Unsupported formats: {', '.join(sorted(unknown))}")

    args.out.mkdir(parents=True, exist_ok=True)
    rows = SCENARIOS[args.scenario](args.rows, Random(args.seed))
    stem = f"demo_{args.scenario.replace('-', '_')}"
    written: list[str] = []
    company_info = {
        "company_name": args.company_name,
        "director": args.director_name,
        "note": DEFAULT_COMPANY_INFO["note"],
    }

    if "csv" in formats:
        path = args.out / f"{stem}.csv"
        write_csv(path, rows)
        written.append(str(path))
        if args.include_company_info:
            company_path = args.out / "company_info.csv"
            write_company_csv(company_path, company_info)
            written.append(str(company_path))
    if "json" in formats:
        path = args.out / f"{stem}.json"
        write_json(path, rows)
        written.append(str(path))
        if args.include_company_info:
            company_path = args.out / "company_info.json"
            write_company_json(company_path, company_info)
            written.append(str(company_path))
    if "xlsx" in formats:
        path = args.out / f"{stem}.xlsx"
        if write_xlsx(path, rows, company_info if args.include_company_info else None, args.logo):
            written.append(str(path))
        else:
            print("Skipped XLSX: install openpyxl to enable Excel export.")

    print(f"Generated {len(rows)} rows for scenario '{args.scenario}'.")
    for path in written:
        print(f"- {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
